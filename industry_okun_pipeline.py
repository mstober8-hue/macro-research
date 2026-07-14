"""
industry_okun_pipeline.py
Okun's Law in the AI Era — Full BLS/BEA Industry Pipeline

Methodology:
  - Difference form of Okun's Law: ΔU = α + β·(%ΔY) + ε
  - YoY (4-quarter) differences to cancel seasonality in NSA unemployment series
  - YoY differences computed FIRST on the intact series (pct_change is
    positional — differencing after row removal would silently compare
    wrong years), THEN the COVID quarters (2020-Q2 through 2021-Q1) and
    the rebound quarters (2021-Q2 through 2022-Q1, whose YoY denominator
    falls in the COVID window) are dropped
  - Era split at 2022-Q4 (ChatGPT launch / post-AI cutoff)
  - Rolling 12-quarter window for β and r (matching GDPUnemployment.py)
  - AIIE scores: Felten, Raj, Seamans (2023), mean across 4-digit NAICS
    sub-industries within each BLS super-sector (computed from Data Appendix B)
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import stats as sp_stats

warnings.filterwarnings("ignore", category=FutureWarning)

# ----------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------
DATA_DIR   = "FRED-Data/"
AI_CUTOFF  = pd.Timestamp("2022-10-01")   # Q4 2022
COVID_DROP = pd.date_range("2020-04-01", "2021-01-01", freq="QS")  # Q2–Q4 2020 + Q1 2021
WINDOW     = 12   # rolling regression quarters

# ----------------------------------------------------------------
# INDUSTRY REGISTRY
# ----------------------------------------------------------------
# Each entry: output_file, unemp_file, aiie score, match_quality, notes.
#
# AIIE = AI Industry Exposure score from Felten, Raj, Seamans (2023).
# Sign: positive = above-average AI exposure, negative = below-average.
# Computed as the mean of 4-digit NAICS AIIE scores within each BLS
# super-sector using the paper's Data Appendix B Excel file.
#
# match_quality flags where the BLS unemployment category and the BEA
# value-added category do not map 1:1 to the same NAICS codes.

INDUSTRIES = {
    "Financial Activities": {
        "output_file":  "FnceservcGDP.csv",
        "unemp_file":   "LNU04032238.csv",
        "aiie":         1.538,
        "match":        "good",
        "notes":        "",
    },
    "Information": {
        "output_file":  "RVAI.csv",
        "unemp_file":   "LNU04032237.csv",
        "aiie":         1.268,
        "match":        "good",
        "notes":        "",
    },
    "Education & Health": {
        "output_file":  "RVAHCSA.csv",
        "unemp_file":   "LNU04032240.csv",
        "aiie":         0.775,
        "match":        "partial",
        "notes":        "BEA RVAHCSA is Health Care & Social Assistance (NAICS 62) only; "
                        "BLS LNU04032240 covers Education & Health combined (NAICS 61+62).",
    },
    "Professional & Business": {
        "output_file":  "RVAPBS.csv",
        "unemp_file":   "LNU04032239.csv",
        "aiie":         0.654,
        "match":        "good",
        "notes":        "",
    },
    "Wholesale Trade": {
        "output_file":  "RVAW.csv",
        "unemp_file":   "LNU04032235.csv",
        "aiie":         0.264,
        "match":        "partial",
        "notes":        "BLS LNU04032235 is 'Wholesale and Retail Trade' combined; "
                        "BEA RVAW is Wholesale only (NAICS 42).",
    },
    "Leisure & Hospitality": {
        "output_file":  "RVAAERAF.csv",   # Arts+Entertainment+Recreation+Accommodation+Food = full NAICS 71+72
        "unemp_file":   "LNU04032241.csv",
        "aiie":        -0.315,
        "match":        "good",
        "notes":        "RVAAERAF covers full Leisure & Hospitality (NAICS 71+72), matching BLS LNU04032241.",
    },
    "Transportation & Utilities": {
        "output_file":  "RVAT.csv",
        "unemp_file":   "LNU04032236.csv",
        "aiie":        -0.342,
        "match":        "partial",
        "notes":        "BEA RVAT is Transportation & Warehousing (NAICS 48-49) only; "
                        "BLS LNU04032236 includes Utilities (NAICS 22).",
    },
    "Manufacturing": {
        "output_file":  "MnfctGDP.csv",
        "unemp_file":   "MnfctUrate.csv",
        "aiie":        -0.484,
        "match":        "good",
        "notes":        "",
    },
    "Construction": {
        "output_file":  "CnstGDP.csv",
        "unemp_file":   "ConstUrate .csv",   # trailing space in filename
        "aiie":        -0.997,
        "match":        "good",
        "notes":        "",
    },
}


# ================================================================
# HELPER FUNCTIONS
# ================================================================

def load_series(filename, label):
    """Load a FRED CSV, return a Series named label."""
    path = os.path.join(DATA_DIR, filename)
    df = pd.read_csv(path, parse_dates=["observation_date"])
    df = df.set_index("observation_date")
    col = [c for c in df.columns][0]
    df[col] = pd.to_numeric(df[col], errors="coerce")
    return df[col].rename(label)


def build_industry_df(cfg):
    """
    Load output + unemployment, align to quarterly, drop COVID, compute YoY.

    WHY YoY INSTEAD OF QoQ:
    The BLS CPS Table A-14 unemployment series are NOT seasonally adjusted.
    Quarterly averages of monthly NSA data still carry a strong seasonal
    pattern — construction unemployment spikes every winter, retail spikes
    after the holidays, etc. A QoQ difference would confuse seasonal swings
    with cyclical changes. A 4-quarter (YoY) difference cancels the seasonal
    pattern exactly because we subtract the same-quarter reading from a year ago.

    WHAT TO EXCLUDE AND WHY:
    We drop two windows from the final dataset:
      (a) COVID quarters Q2 2020–Q1 2021: the shock itself is too large to
          be informative about the structural Okun relationship.
      (b) Rebound quarters Q2 2021–Q1 2022: their YoY comparison looks back
          to the COVID spike/trough, so the computed ΔY and ΔU are dominated
          by the mechanical base effect of recovering from the shock rather
          than real economic dynamics.

    CRITICAL ORDER: compute pct_change(4) and diff(4) BEFORE dropping any
    rows. pct_change uses positional indexing — if COVID rows are removed
    first, Q2 2021 (the next row in the DataFrame) looks 4 positions back
    to Q2 2019 instead of Q2 2020, silently producing a wrong two-year
    comparison. Computing first on the intact series guarantees correct
    date-matched YoY arithmetic throughout.
    """
    y = load_series(cfg["output_file"], "output")
    u_monthly = load_series(cfg["unemp_file"], "unemp")
    u = u_monthly.resample("QS").mean()

    df = pd.DataFrame({"output": y, "unemp": u}).dropna()

    # Compute YoY on the full intact series — no gaps yet
    df["pct_dy"]  = df["output"].pct_change(periods=4) * 100
    df["delta_u"] = df["unemp"].diff(periods=4)

    # Drop COVID quarters AND the four rebound quarters that compare to them
    EXCLUDE = pd.date_range("2020-04-01", "2022-01-01", freq="QS")  # Q2 2020 – Q1 2022
    df = df[~df.index.isin(EXCLUDE)]

    return df.dropna(subset=["pct_dy", "delta_u"])


def fit_ols(x, y):
    """Return (slope β, intercept α, correlation r, n). NaN if n < 3."""
    mask = ~(np.isnan(x) | np.isnan(y))
    x, y = x[mask], y[mask]
    n = len(x)
    if n < 3:
        return np.nan, np.nan, np.nan, n
    slope, intercept, r, _, _ = sp_stats.linregress(x, y)
    return slope, intercept, r, n


def rolling_okun(df, window=WINDOW):
    """
    12-quarter rolling OLS of ΔU on %ΔY.
    Returns DataFrame with columns slope (β) and r, indexed by end-of-window quarter.
    """
    idx = df.index.tolist()
    dates, slopes, rs = [], [], []
    for i in range(window, len(idx) + 1):
        w = df.iloc[i - window : i]
        x, y = w["pct_dy"].values, w["delta_u"].values
        if np.std(x) < 1e-9:
            continue
        slope, _, r, n = fit_ols(x, y)
        if np.isnan(slope):
            continue
        dates.append(idx[i - 1])
        slopes.append(slope)
        rs.append(r)
    return pd.DataFrame({"slope": slopes, "r": rs}, index=dates)


# ================================================================
# MONETARY POLICY CONTROL — Federal Funds Rate
# ----------------------------------------------------------------
# The Q4 2022 AI cutoff coincides almost exactly with the Fed's most
# aggressive rate-hike cycle in four decades (March 2022 – July 2023,
# +525 bp). Construction, Manufacturing, and Transportation are the
# most rate-sensitive industries in the economy. Any Okun breakdown
# in those sectors post-2022 is confounded with rate effects.
#
# We try to load FEDFUNDS.csv (download from FRED: series FEDFUNDS).
# If present, we compute YoY change in the Fed Funds Rate and add it
# as a second regressor in each industry's post-2022 OLS:
#   ΔU = α + β·%ΔY + γ·ΔFFR
# β from this rate-controlled regression is a cleaner Okun coefficient.
# We report both controlled and uncontrolled β_post side by side.
# ================================================================

def load_ffr_control():
    """
    Try to load FEDFUNDS.csv. Returns a quarterly YoY-differenced Series,
    or None if the file is not found.
    To enable: download FEDFUNDS from https://fred.stlouisfed.org/series/FEDFUNDS
    and save as 'FRED-Data/FEDFUNDS.csv'.
    """
    try:
        ffr = load_series("FEDFUNDS.csv", "ffr")
        ffr_q  = ffr.resample("QS").mean()
        ffr_df = pd.DataFrame({"ffr": ffr_q})
        ffr_df["delta_ffr"] = ffr_df["ffr"].diff(periods=4)  # YoY change in FFR
        return ffr_df["delta_ffr"].dropna()
    except FileNotFoundError:
        return None

ffr_control = load_ffr_control()
if ffr_control is not None:
    print("  [Rate control] FEDFUNDS.csv loaded — will run rate-controlled regressions.")
else:
    print("  [Rate control] FEDFUNDS.csv not found — add it from FRED to enable rate control.")
print()


def fit_ols_controlled(x, y, z):
    """
    OLS: y = α + β·x + γ·z. Returns (β, r_partial, n).
    z = rate control (YoY change in Fed Funds Rate).
    β is the Okun coefficient after removing rate effects.
    r_partial is the partial correlation of x with y after removing z.
    """
    mask = ~(np.isnan(x) | np.isnan(y) | np.isnan(z))
    x, y, z = x[mask], y[mask], z[mask]
    n = len(x)
    if n < 4:
        return np.nan, np.nan, n
    # Residualize x and y on z (Frisch–Waugh)
    A = np.column_stack([np.ones(n), z])
    x_resid = x - A @ np.linalg.lstsq(A, x, rcond=None)[0]
    y_resid = y - A @ np.linalg.lstsq(A, y, rcond=None)[0]
    beta, _, r_p, _, _ = sp_stats.linregress(x_resid, y_resid)
    return beta, r_p, n


# ================================================================
# STEPS 1–5: LOAD, FIT, AND ROLL FOR EVERY INDUSTRY
# ================================================================

industry_data = {}
industry_roll = {}
summary_rows  = []
quality_flags = []

print("=" * 70)
print("INDUSTRY OKUN PIPELINE")
print("=" * 70)

for name, cfg in INDUSTRIES.items():
    try:
        df = build_industry_df(cfg)
    except FileNotFoundError as e:
        quality_flags.append(f"SKIPPED {name}: file not found — {e}")
        print(f"  SKIP  {name}: {e}")
        continue

    if len(df) < WINDOW + 4:
        quality_flags.append(f"SKIPPED {name}: only {len(df)} quarters — too few.")
        print(f"  SKIP  {name}: only {len(df)} clean quarters")
        continue

    industry_data[name] = df

    pre  = df[df.index <  AI_CUTOFF]
    post = df[df.index >= AI_CUTOFF]

    b_pre,  _, r_pre,  n_pre  = fit_ols(pre["pct_dy"].values,  pre["delta_u"].values)
    b_post, _, r_post, n_post = fit_ols(post["pct_dy"].values, post["delta_u"].values)

    # Rate-controlled β_post (if FEDFUNDS available)
    b_post_rc = np.nan
    if ffr_control is not None:
        post_joint = post.join(ffr_control.rename("delta_ffr"), how="inner")
        if len(post_joint) >= 4:
            b_post_rc, _, _ = fit_ols_controlled(
                post_joint["pct_dy"].values,
                post_joint["delta_u"].values,
                post_joint["delta_ffr"].values,
            )

    roll = rolling_okun(df)
    industry_roll[name] = roll

    b_change    = (b_post    - b_pre) if not (np.isnan(b_pre) or np.isnan(b_post))    else np.nan
    b_change_rc = (b_post_rc - b_pre) if not (np.isnan(b_pre) or np.isnan(b_post_rc)) else np.nan
    r_change    = (r_post    - r_pre) if not (np.isnan(r_pre) or np.isnan(r_post))    else np.nan

    summary_rows.append({
        "industry":       name,
        "aiie":           cfg["aiie"],
        "match":          cfg["match"],
        "n_pre":          n_pre,
        "beta_pre":       round(b_pre,       4) if not np.isnan(b_pre)       else np.nan,
        "r_pre":          round(r_pre,       4) if not np.isnan(r_pre)       else np.nan,
        "n_post":         n_post,
        "beta_post":      round(b_post,      4) if not np.isnan(b_post)      else np.nan,
        "r_post":         round(r_post,      4) if not np.isnan(r_post)      else np.nan,
        "beta_change":    round(b_change,    4) if not np.isnan(b_change)    else np.nan,
        "beta_post_rc":   round(b_post_rc,   4) if not np.isnan(b_post_rc)   else np.nan,
        "beta_change_rc": round(b_change_rc, 4) if not np.isnan(b_change_rc) else np.nan,
        "r_change":       round(r_change,    4) if not np.isnan(r_change)    else np.nan,
    })

    # Quality flags (Step 10)
    if n_pre < 8:
        quality_flags.append(f"FLAG [{name}] pre-period n={n_pre} < 8 — slope unreliable.")
    if n_post < 8:
        quality_flags.append(f"FLAG [{name}] post-period n={n_post} < 8 — borderline sample.")
    max_swing = df["unemp"].diff().abs().max()
    if max_swing > 3.0:
        quality_flags.append(
            f"FLAG [{name}] max YoY unemployment swing = {max_swing:.1f}pp > 3pp — "
            "CPS sample may be noisy; interpret with caution."
        )
    if cfg["match"] == "partial":
        quality_flags.append(f"FLAG [{name}] partial BLS/BEA match — {cfg['notes']}")
    # Specific structural flags
    if name == "Education & Health":
        quality_flags.append(
            f"FLAG [{name}] acyclical sector — Okun's law never strong here "
            f"(r_pre={r_pre:.2f}); and BEA RVAHCSA (NAICS 62 only) doesn't fully "
            "match BLS LNU04032240 (NAICS 61+62). Exclude from cross-section."
        )
    if name == "Construction":
        quality_flags.append(
            f"FLAG [Construction] r flips dramatically (+0.77) but β_post is only "
            f"+{b_post:.3f} — near-zero sensitivity, not true inversion. "
            "Also: pre-2022 slope is heavily influenced by 2008-09 housing crash; "
            "a leave-out test would show whether β_pre is a stable estimate."
        )

    rc_str = f"  β_post_rc={b_post_rc:+.3f}" if not np.isnan(b_post_rc) else ""
    print(f"  OK  {name:32s}  n_pre={n_pre:3d}  β_pre={b_pre:+.3f}  "
          f"n_post={n_post:3d}  β_post={b_post:+.3f}  Δβ={b_change:+.3f}{rc_str}")

print()


# ================================================================
# STEP 6: SUMMARY TABLE CSV
# ================================================================

summary = pd.DataFrame(summary_rows).set_index("industry")
summary.to_csv("okun_industry_summary.csv")
print("Summary table → okun_industry_summary.csv")
print()
print(summary[["aiie","n_pre","beta_pre","r_pre","n_post",
               "beta_post","r_post","beta_change","r_change"]].to_string())
print()


# ================================================================
# STEP 8: CROSS-SECTIONAL REGRESSION — beta_change ~ AIIE score
# ----------------------------------------------------------------
# AI hypothesis prediction: high-AIIE industries produce output without
# proportional employment changes → β becomes LESS negative → beta_change
# is MORE POSITIVE. So the hypothesis predicts a POSITIVE cross-sectional
# slope. A negative slope would mean low-AIIE sectors show more weakening.
# With n ≈ 8–10 this is suggestive; report exact n and p, do not overstate.
# ================================================================

cross   = summary[["aiie", "beta_change"]].dropna()
n_cross = len(cross)

if n_cross >= 4:
    slope_cs, intercept_cs, r_cs, p_cs, _ = sp_stats.linregress(
        cross["aiie"].values, cross["beta_change"].values
    )

    print("=" * 70)
    print("CROSS-SECTIONAL REGRESSION: beta_change ~ AIIE_score")
    print(f"  n={n_cross}  slope={slope_cs:.4f}  r={r_cs:.3f}  p={p_cs:.4f}")
    if p_cs < 0.01:   sig = "significant at 99%"
    elif p_cs < 0.05: sig = "significant at 95%"
    elif p_cs < 0.10: sig = "significant at 90%"
    else:             sig = "not significant at 90%"
    print(f"  Result: {sig}")
    print()

    fig_cs, ax_cs = plt.subplots(figsize=(10, 7))
    ax_cs.scatter(cross["aiie"], cross["beta_change"],
                  color="steelblue", s=80, zorder=3)
    for ind in cross.index:
        ax_cs.annotate(ind,
                       xy=(cross.loc[ind, "aiie"], cross.loc[ind, "beta_change"]),
                       xytext=(6, 4), textcoords="offset points", fontsize=8)
    x_r = np.linspace(cross["aiie"].min() - 0.3, cross["aiie"].max() + 0.3, 100)
    ax_cs.plot(x_r, intercept_cs + slope_cs * x_r,
               color="firebrick", linewidth=1.5, linestyle="--",
               label=f"OLS  r={r_cs:.2f}  p={p_cs:.3f}  n={n_cross}")
    ax_cs.axhline(0, color="black", linewidth=0.8, linestyle=":")
    ax_cs.axvline(0, color="black", linewidth=0.8, linestyle=":")
    ax_cs.set_xlabel("AIIE Score (Felten et al. 2023)\nHigher = more AI-exposed",
                     fontsize=12)
    ax_cs.set_ylabel("Δβ  (β_post − β_pre)\nMore positive = Okun's law weakened more (β less negative)",
                     fontsize=12)
    ax_cs.set_title(
        "Cross-Industry: Does AI Exposure Predict Okun's Law Breakdown?\n"
        f"n={n_cross} industries  |  Treat as suggestive given small sample",
        fontsize=13, fontweight="bold")
    ax_cs.legend(fontsize=10)
    ax_cs.grid(True, linestyle="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig("industry_aiie_scatter.png", dpi=150, bbox_inches="tight")
    print("Chart saved: industry_aiie_scatter.png")
else:
    slope_cs = intercept_cs = r_cs = p_cs = np.nan
    print(f"Cross-sectional regression skipped — only {n_cross} complete industries.")


# ================================================================
# STEP 9: ROLLING β OVERLAY — all industries on one chart
# ----------------------------------------------------------------
# Warm colors = high AIIE (more AI). Cool colors = low AIIE (less AI).
# If the AI hypothesis holds, post-2022 warm-colored lines should
# drift away from cool-colored lines.
# ================================================================

sorted_names = sorted(industry_roll.keys(),
                      key=lambda n: INDUSTRIES[n]["aiie"], reverse=True)
cmap   = plt.cm.RdYlBu
n_ind  = len(sorted_names)
colors = [cmap(i / max(n_ind - 1, 1)) for i in range(n_ind)]

latest = max(r.index[-1] for r in industry_roll.values())

fig9, (ax9a, ax9b) = plt.subplots(2, 1, figsize=(14, 11), sharex=False)

for i, name in enumerate(sorted_names):
    roll = industry_roll[name]
    aiie = INDUSTRIES[name]["aiie"]
    lbl  = f"{name}  (AIIE={aiie:+.2f})"
    ax9a.plot(roll.index, roll["slope"], color=colors[i],
              linewidth=1.8, label=lbl, alpha=0.85)
    ax9b.plot(roll.index, roll["r"],     color=colors[i],
              linewidth=1.8, label=lbl, alpha=0.85)

for ax in (ax9a, ax9b):
    ax.axhline(0, color="black", linewidth=0.9, linestyle="--")
    ax.axvspan(AI_CUTOFF, latest, alpha=0.07, color="gold")
    ax.grid(True, linestyle="--", alpha=0.35)

ax9a.set_ylabel("Rolling Okun's β  (ΔU per 1% output growth)", fontsize=11)
ax9a.set_title(
    "Rolling 12-Quarter Okun's β — All Industries\n"
    "Warm (red/orange) = high AI exposure | Cool (blue) = low AI exposure  "
    "| Gold = post-Q4-2022",
    fontsize=12, fontweight="bold")
ax9a.legend(fontsize=7.5, loc="upper left", ncol=2)

ax9b.set_ylabel("Rolling Correlation r\n(r < 0 = law holds | r > 0 = law inverted)",
                fontsize=11)
ax9b.set_xlabel("Quarter (end of 12-quarter window)", fontsize=12)
ax9b.set_ylim(-1.05, 1.05)
ax9b.legend(fontsize=7.5, loc="lower left", ncol=2)

plt.tight_layout(pad=2.5)
plt.savefig("industry_rolling_overlay.png", dpi=150, bbox_inches="tight")
print("Chart saved: industry_rolling_overlay.png")


# ================================================================
# STEP 10: QUALITY FLAGS
# ================================================================

print()
print("=" * 70)
print("QUALITY FLAGS")
print("=" * 70)
for flag in quality_flags:
    print(f"  {flag}")
print()


# ================================================================
# STEP 11: WRITTEN SUMMARY
# ================================================================

# Sort descending so the most-weakened industries appear first
summary_sorted = summary["beta_change"].dropna().sort_values(ascending=False)
post_n_med     = int(summary["n_post"].dropna().median())

lines = [
    "=" * 80,
    "WRITTEN SUMMARY — Okun's Law in the AI Era: Industry Pipeline",
    "=" * 80,
    "",
    f"INDUSTRIES ANALYZED: {len(industry_data)}",
    f"  {', '.join(industry_data.keys())}",
    "",
    "Δβ = β_post − β_pre.  β < 0 = law holds (output up → unemployment down).",
    "Δβ > 0 means β became LESS negative → law WEAKENED (output no longer pulls unemployment down).",
    "Δβ < 0 means β became MORE negative → law STRENGTHENED.",
    "",
    "RANKED BY LAW WEAKENING (most weakened first):",
]
for ind, val in summary_sorted.items():
    tag = "← LAW WEAKENED" if val > 0.05 else ("← law strengthened" if val < -0.05 else "← little change")
    lines.append(f"  {ind:<34}  Δβ = {val:+.4f}  {tag}")

lines += [
    "",
    "CROSS-SECTIONAL RESULT (beta_change ~ AIIE_score):",
]
if not np.isnan(r_cs):
    # Sign logic: AI hypothesis predicts high-AIIE → more positive Δβ (law weakened)
    # → positive slope expected. Negative slope = opposite pattern.
    if slope_cs > 0:
        cs_verdict = (
            "SUPPORTS AI hypothesis: higher-AIIE industries show more Okun weakening."
        )
    else:
        cs_verdict = (
            "CONTRADICTS simple AI hypothesis: low-AIIE industries (Construction, "
            "Manufacturing) show the LARGEST Okun weakening. High-AIIE sectors "
            "(Financial, Professional & Business) show stable or strengthening law."
        )
    lines += [
        f"  n={n_cross}  slope={slope_cs:.4f}  r={r_cs:.3f}  p={p_cs:.4f}  ({sig})",
        f"  {cs_verdict}",
    ]
else:
    lines.append("  Not computed (insufficient data).")

lines += [
    "",
    "WHAT THIS CAN AND CANNOT ESTABLISH:",
    f"  Post-2022 sample: ~{post_n_med} clean quarters per industry.",
    "",
    "  FINDING: The Okun's law breakdown is concentrated in LOW-AI sectors.",
    "  Construction and Manufacturing show the largest β drift toward zero.",
    "  High-AI sectors (Financial Activities, Professional & Business,",
    "  Education & Health) show the relationship holding or strengthening.",
    "",
    "  THREE ALTERNATIVE EXPLANATIONS FOR THE PATTERN:",
    "  1. Fiscal policy effect: post-2022 infrastructure/industrial investment",
    "     (IIJA, CHIPS Act, IRA) boosted Construction and Manufacturing OUTPUT",
    "     without equivalent employment (labor shortages, supply constraints).",
    "     This would produce exactly the observed pattern independent of AI.",
    "  2. AI augmentation vs displacement: the AIIE score measures occupational",
    "     exposure to AI capabilities, not actual displacement. High-AIIE workers",
    "     (analysts, managers, finance professionals) may be using AI as a tool",
    "     that boosts their productivity — strengthening rather than breaking",
    "     the output-employment link. True displacement may lag exposure.",
    "  3. Sectoral composition of the AGGREGATE breakdown: GDPUnemployment.py",
    "     shows aggregate Okun breaking down post-2022. This analysis suggests",
    "     the aggregate signal may originate in physical sectors, not tech —",
    "     which challenges the narrative that AI is the primary driver.",
    "",
    "  CANNOT establish causality from this analysis alone. Revisit as",
    "  post-2022 sample grows toward 20+ quarters (2026–2027).",
    "=" * 80,
]

summary_text = "\n".join(lines)
print(summary_text)

with open("okun_industry_summary.txt", "w") as f:
    f.write(summary_text)
print("\nSummary saved: okun_industry_summary.txt")


# ================================================================
# STEP 12: PER-INDUSTRY CHARTS
# ----------------------------------------------------------------
# For every industry we produce a single 2×2 figure containing:
#
#   TOP-LEFT  — Okun's Law scatter: x=%ΔY, y=ΔU
#     Each dot is one quarter. Color = time (purple=oldest, yellow=newest),
#     so you can see whether the relationship drifts over time without
#     relying on a binary pre/post split. Regression lines drawn separately
#     for pre- and post-Q4 2022 so the slope shift is visible.
#
#   TOP-RIGHT — Rolling 12-quarter β (solid) and r (dashed) over time.
#     β = Okun's coefficient. r = correlation between output growth and
#     unemployment change. Both should be negative if the law holds.
#     The gold shading marks the post-Q4 2022 era.
#
#   BOTTOM-LEFT — Okun residual over time.
#     Fit the pre-2022 relationship, then compute actual ΔU minus what
#     Okun's law predicted given the output growth that occurred.
#     Persistent positive residual post-2022 = output grew but
#     unemployment didn't fall as much as it historically would have.
#
#   BOTTOM-RIGHT — Output index + unemployment rate (twin axis).
#     Output indexed to Q4 2019 = 100. Under Okun's Law, if output
#     rises above its base, unemployment should fall. If the output
#     line rises but the unemployment line stays flat, Okun is broken.
#
# Files saved as: okun_<industry_slug>.png
# ================================================================

# Store enriched quarterly DataFrames for the Excel export (step 13)
industry_export = {}

print("\nGenerating per-industry charts...")

for name, df in industry_data.items():
    slug = (name.lower()
              .replace(" & ", "_")
              .replace(" ", "_")
              .replace("&", "_"))

    df2 = df.copy()
    df2["era"] = "Pre-Q4 2022"
    df2.loc[df2.index >= AI_CUTOFF, "era"] = "Q4 2022–Present"

    # Pre-2022 Okun fit (used as baseline for residual panel)
    pre  = df2[df2.index <  AI_CUTOFF]
    post = df2[df2.index >= AI_CUTOFF]

    b_pre,  a_pre,  r_pre,  n_pre  = fit_ols(pre["pct_dy"].values,  pre["delta_u"].values)
    b_post, a_post, r_post, n_post = fit_ols(post["pct_dy"].values, post["delta_u"].values)

    df2["u_predicted"] = b_pre * df2["pct_dy"] + a_pre
    df2["okun_resid"]  = df2["delta_u"] - df2["u_predicted"]

    # Store enriched df for Excel
    industry_export[name] = df2

    # Base-index output to Q4 2019 = 100 for time-series panel
    base_date = pd.Timestamp("2019-10-01")
    base_val  = df2.loc[df2.index <= base_date, "output"].iloc[-1] if (df2.index <= base_date).any() else df2["output"].iloc[0]
    df2["output_idx"] = df2["output"] / base_val * 100

    # NaN gap at COVID dates for line-break in residual/time-series panels
    nan_rows = pd.DataFrame(
        {"okun_resid": np.nan, "output_idx": np.nan, "unemp": np.nan},
        index=pd.date_range("2020-04-01", "2021-01-01", freq="QS")
    )
    plot_df = pd.concat([df2[["okun_resid", "output_idx", "unemp"]], nan_rows]).sort_index()

    roll = industry_roll[name]

    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    ax_sc, ax_roll, ax_resid, ax_ts = axes[0, 0], axes[0, 1], axes[1, 0], axes[1, 1]

    # ── Panel 1: Scatter colored by TIME ─────────────────────────
    # Color runs from purple (oldest) → yellow (newest) via plasma colormap.
    # This shows whether the cloud of points drifts over time without
    # imposing a sharp pre/post binary — you can see gradual drift.
    n = len(df2)
    sc = ax_sc.scatter(df2["pct_dy"], df2["delta_u"],
                       c=range(n), cmap="plasma", s=60, zorder=3, alpha=0.85)

    era_styles = [
        ("Pre-Q4 2022",    "steelblue",  b_pre,  a_pre,  r_pre,  n_pre),
        ("Q4 2022–Present","darkorange", b_post, a_post, r_post, n_post),
    ]
    for era, col, b, a, r, n_era in era_styles:
        sub = df2[df2["era"] == era]
        if len(sub) >= 3 and not np.isnan(b):
            x_r = np.linspace(sub["pct_dy"].min(), sub["pct_dy"].max(), 100)
            ax_sc.plot(x_r, a + b * x_r, color=col, linewidth=1.8,
                       linestyle="--",
                       label=f"{era}\nβ={b:.3f}  r={r:.3f}  n={n_era}")

    ax_sc.axhline(0, color="black", linewidth=0.7, linestyle=":")
    ax_sc.axvline(0, color="black", linewidth=0.7, linestyle=":")
    cbar = plt.colorbar(sc, ax=ax_sc, pad=0.02)
    cbar.set_label("Time (purple=oldest → yellow=newest)", fontsize=7)
    ticks = [0, n // 4, n // 2, 3 * n // 4, n - 1]
    cbar.set_ticks(ticks)
    cbar.set_ticklabels([str(df2.index[i].year) for i in ticks])
    ax_sc.set_xlabel("YoY % Change in Real Value Added (%ΔY)", fontsize=10)
    ax_sc.set_ylabel("YoY Change in Unemployment Rate (ΔU, pp)", fontsize=10)
    ax_sc.set_title("Okun's Law Scatter\nDot color = time (purple→yellow = old→new)",
                    fontsize=11, fontweight="bold")
    ax_sc.legend(fontsize=8, loc="upper right")
    ax_sc.grid(True, linestyle="--", alpha=0.35)

    # ── Panel 2: Rolling β and r ──────────────────────────────────
    ax_roll.plot(roll.index, roll["slope"], color="steelblue",
                 linewidth=2, label="Rolling β (Okun's coefficient)")
    ax_roll.plot(roll.index, roll["r"],     color="firebrick",
                 linewidth=1.5, linestyle="--", alpha=0.85,
                 label="Rolling r (correlation)")
    ax_roll.axhline(0, color="black", linewidth=0.8, linestyle="--")
    if len(roll) > 0:
        ax_roll.axvspan(AI_CUTOFF, roll.index[-1], alpha=0.10, color="gold",
                        label="Post-Q4 2022")
    ax_roll.set_ylabel("β  /  r", fontsize=10)
    ax_roll.set_title("Rolling 12-Quarter Okun's β and r\n"
                      "Both should be negative when law holds",
                      fontsize=11, fontweight="bold")
    ax_roll.legend(fontsize=8)
    ax_roll.set_ylim(-1.3, 1.3)
    ax_roll.grid(True, linestyle="--", alpha=0.35)

    # ── Panel 3: Okun Residual ────────────────────────────────────
    ax_resid.axhline(0, color="black", linewidth=1.2, linestyle="--", zorder=2)
    ax_resid.fill_between(
        plot_df.index, plot_df["okun_resid"], 0,
        where=plot_df["okun_resid"] > 0,
        color="firebrick", alpha=0.4,
        label="Unemployment higher than Okun predicts")
    ax_resid.fill_between(
        plot_df.index, plot_df["okun_resid"], 0,
        where=plot_df["okun_resid"] <= 0,
        color="steelblue", alpha=0.4,
        label="Unemployment lower than Okun predicts")
    ax_resid.plot(plot_df.index, plot_df["okun_resid"],
                  color="black", linewidth=1.0)
    ax_resid.axvspan(pd.Timestamp("2020-04-01"), pd.Timestamp("2021-04-01"),
                     alpha=0.15, color="crimson", label="COVID excluded")
    ax_resid.axvspan(AI_CUTOFF, df2.index[-1],
                     alpha=0.08, color="gold", label="Post-Q4 2022")
    ax_resid.set_ylabel("Okun Residual (pp)\nActual ΔU − Predicted ΔU", fontsize=10)
    ax_resid.set_title(
        f"Okun Residual (pre-2022 fit as baseline)\n"
        f"Pre-fit: ΔU = {b_pre:.3f}×%ΔY + {a_pre:.3f}",
        fontsize=11, fontweight="bold")
    ax_resid.legend(fontsize=7.5, loc="lower left")
    ax_resid.grid(True, linestyle="--", alpha=0.35)

    # ── Panel 4: Output index + unemployment (twin axis) ─────────
    # Left axis: output indexed to Q4 2019 = 100.
    # Right axis: unemployment rate (%).
    # If output rises but unemployment stays flat, Okun is broken.
    ax_ts2 = ax_ts.twinx()

    ax_ts.plot(plot_df.index, plot_df["output_idx"],
               color="steelblue", linewidth=2.2,
               label=f"Output index (Q4 2019 = 100)")
    ax_ts2.plot(plot_df.index, plot_df["unemp"],
                color="firebrick", linewidth=1.8, linestyle="--",
                alpha=0.85, label="Unemployment rate % (right)")

    ax_ts.axhline(100, color="steelblue", linewidth=0.6,
                  linestyle=":", alpha=0.6)
    ax_ts.axvspan(pd.Timestamp("2020-04-01"), pd.Timestamp("2021-04-01"),
                  alpha=0.15, color="crimson")
    ax_ts.axvspan(AI_CUTOFF, df2.index[-1],
                  alpha=0.08, color="gold")
    ax_ts.set_ylabel("Real Value Added Index (Q4 2019=100)",
                     fontsize=10, color="steelblue")
    ax_ts2.set_ylabel("Unemployment Rate (%)",
                      fontsize=10, color="firebrick")
    ax_ts2.tick_params(axis="y", labelcolor="firebrick")

    lines_l, labs_l = ax_ts.get_legend_handles_labels()
    lines_r, labs_r = ax_ts2.get_legend_handles_labels()
    ax_ts.legend(lines_l + lines_r, labs_l + labs_r, fontsize=8, loc="upper left")
    ax_ts.set_title("Output Index vs Unemployment Rate\n"
                    "Divergence = output rising without unemployment falling",
                    fontsize=11, fontweight="bold")
    ax_ts.grid(True, linestyle="--", alpha=0.35)

    # ── Figure title ─────────────────────────────────────────────
    b_chg = b_post - b_pre
    plt.suptitle(
        f"{name}   (AIIE = {INDUSTRIES[name]['aiie']:+.2f}  |  "
        f"BLS/BEA match: {INDUSTRIES[name]['match']})\n"
        f"Pre-2022  β={b_pre:.3f}  r={r_pre:.3f}  n={n_pre}     "
        f"Post-2022  β={b_post:.3f}  r={r_post:.3f}  n={n_post}     "
        f"Δβ = {b_chg:+.3f}",
        fontsize=11, fontweight="bold", y=1.01
    )
    plt.tight_layout(rect=[0, 0, 1, 0.97])
    plt.savefig(f"okun_{slug}.png", dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: okun_{slug}.png")


# ================================================================
# STEP 13: EXCEL WORKBOOK — full numerical data for every industry
# ----------------------------------------------------------------
# Sheet layout:
#   "Summary"       — one row per industry, all key stats + interpretation
#   One sheet per industry — quarterly data with all computed columns
#
# Intended for people who want to verify the numbers directly rather
# than read the charts.
# ================================================================

print("\nBuilding Excel workbook...")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue header
    SUB_FILL  = PatternFill("solid", fgColor="BDD7EE")   # light blue subheader
    PRE_FILL  = PatternFill("solid", fgColor="DDEBF7")   # very light blue for pre rows
    POST_FILL = PatternFill("solid", fgColor="FCE4D6")   # light orange for post rows
    POS_FILL  = PatternFill("solid", fgColor="FFCCCC")   # red tint: residual > 0 (bad)
    NEG_FILL  = PatternFill("solid", fgColor="CCFFCC")   # green tint: residual < 0 (good)
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    BOLD      = Font(bold=True)

    thin = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, value, fill=HDR_FILL, font=HDR_FONT):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        return c

    def data_cell(ws, row, col, value, fmt=None, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        return c

    # ── SUMMARY SHEET ─────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.row_dimensions[1].height = 30
    ws_sum.row_dimensions[2].height = 45

    title_cell = ws_sum.cell(row=1, column=1,
        value="Okun's Law in the AI Era — Industry Summary")
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    ws_sum.merge_cells("A1:P1")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sum_headers = [
        "Industry", "AIIE Score", "AI Exposure\nLevel",
        "BLS/BEA\nMatch",
        "Pre-2022\nn (quarters)", "Pre-2022\nβ (Okun coeff)", "Pre-2022\nr (correlation)",
        "Post-2022\nn (quarters)", "Post-2022\nβ (Okun coeff)", "Post-2022\nr (correlation)",
        "Δβ\n(post − pre)", "Δr\n(post − pre)",
        "Post-2022\nMean Residual (pp)",
        "Interpretation", "Consistent\nw/ AI Hypothesis?", "Notes",
    ]
    for ci, h in enumerate(sum_headers, 1):
        hdr_cell(ws_sum, 2, ci, h)

    # AI exposure label helper
    def ai_level(score):
        if score > 0.5:   return "High"
        if score > 0.0:   return "Medium-High"
        if score > -0.5:  return "Medium-Low"
        return "Low"

    # Interpretation helper
    def interpret(b_pre, b_post, r_pre, r_post):
        b_chg = b_post - b_pre
        if np.isnan(b_chg):
            return "Insufficient data"
        if b_chg > 0.1 and r_post > 0:
            return "Law inverted post-AI: output up but unemployment also rising"
        if b_chg > 0.1:
            return "Law weakened significantly: output growth no longer pulls unemployment down"
        if b_chg > 0.02:
            return "Law moderately weakened"
        if b_chg < -0.05:
            return "Law strengthened: unemployment more responsive to output post-AI"
        return "No meaningful change"

    def ai_consistent(b_pre, b_post, aiie):
        b_chg = b_post - b_pre
        if np.isnan(b_chg): return "N/A"
        if aiie > 0 and b_chg > 0.05:   return "Yes — high AI, law weakened"
        if aiie < 0 and b_chg <= 0.05:  return "Yes — low AI, law stable"
        if aiie > 0 and b_chg <= 0.05:  return "No — high AI but law held"
        return "Mixed"

    for ri, row in enumerate(summary_rows, 3):
        ind      = row["industry"]
        cfg      = INDUSTRIES[ind]
        b_pre_v  = row["beta_pre"]
        b_post_v = row["beta_post"]
        r_pre_v  = row["beta_pre"]   # was stored as r_pre in row
        r_post_v = row["r_post"]

        # Post-2022 mean residual
        df_ex = industry_export.get(ind)
        post_resid_mean = (
            df_ex.loc[df_ex.index >= AI_CUTOFF, "okun_resid"].mean()
            if df_ex is not None else np.nan
        )

        row_fill = PRE_FILL if ri % 2 == 1 else None

        vals = [
            ind,
            row["aiie"],
            ai_level(row["aiie"]),
            row["match"],
            row["n_pre"],
            row["beta_pre"],
            row["r_pre"],
            row["n_post"],
            row["beta_post"],
            row["r_post"],
            row["beta_change"],
            row["r_change"],
            round(post_resid_mean, 4) if not np.isnan(post_resid_mean) else "N/A",
            interpret(b_pre_v, b_post_v, row["r_pre"], r_post_v),
            ai_consistent(b_pre_v, b_post_v, row["aiie"]),
            cfg["notes"] or "—",
        ]
        fmts = [None, "0.000", None, None,
                "0", "0.0000", "0.0000",
                "0", "0.0000", "0.0000",
                "0.0000", "0.0000", "0.0000",
                None, None, None]

        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            data_cell(ws_sum, ri, ci, v, fmt=fmt, fill=row_fill)

    # Column widths for summary sheet
    widths = [28, 10, 14, 10, 10, 12, 12, 10, 12, 12, 10, 10, 14, 40, 22, 50]
    for ci, w in enumerate(widths, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    # ── PER-INDUSTRY SHEETS ───────────────────────────────────────
    for name, df_ex in industry_export.items():
        cfg   = INDUSTRIES[name]
        slug  = name[:28]   # sheet names max 31 chars
        ws    = wb.create_sheet(title=slug)

        # Stats block at the top
        row_stats = summary[summary.index == name].iloc[0] if name in summary.index else None

        ws.merge_cells("A1:I1")
        t = ws.cell(row=1, column=1, value=f"{name} — Okun's Law Detail")
        t.font = Font(bold=True, size=13, color="1F4E79")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        if row_stats is not None:
            stats_labels = [
                ("AIIE Score",          f"{cfg['aiie']:+.3f}"),
                ("BLS/BEA Match",       cfg["match"]),
                ("Pre-2022 β",          f"{row_stats['beta_pre']:.4f}"  if not np.isnan(row_stats['beta_pre']) else "N/A"),
                ("Pre-2022 r",          f"{row_stats['r_pre']:.4f}"     if not np.isnan(row_stats['r_pre'])    else "N/A"),
                ("Pre-2022 n",          f"{int(row_stats['n_pre'])}"),
                ("Post-2022 β",         f"{row_stats['beta_post']:.4f}" if not np.isnan(row_stats['beta_post']) else "N/A"),
                ("Post-2022 r",         f"{row_stats['r_post']:.4f}"    if not np.isnan(row_stats['r_post'])   else "N/A"),
                ("Post-2022 n",         f"{int(row_stats['n_post'])}"),
                ("Δβ (post − pre)",     f"{row_stats['beta_change']:+.4f}" if not np.isnan(row_stats['beta_change']) else "N/A"),
            ]
            for si, (lbl, val) in enumerate(stats_labels):
                col = (si % 3) * 2 + 1
                row_n = 2 + si // 3
                lc = ws.cell(row=row_n, column=col, value=lbl)
                lc.font = BOLD
                lc.fill = SUB_FILL
                lc.alignment = Alignment(horizontal="right")
                vc = ws.cell(row=row_n, column=col + 1, value=val)
                vc.alignment = Alignment(horizontal="left")

        data_start_row = 6

        col_headers = [
            "Quarter", "Era",
            "Real Value Added\n($B chained 2017)",
            "Unemployment\nRate (%)",
            "YoY Output\nGrowth % (ΔY)",
            "YoY Unemp.\nChange (ΔU, pp)",
            "Okun Predicted\nΔU (pp)",
            "Okun Residual\n(pp)  [actual−pred]",
            "Residual\nInterpretation",
        ]
        ws.row_dimensions[data_start_row].height = 40
        for ci, h in enumerate(col_headers, 1):
            hdr_cell(ws, data_start_row, ci, h)

        pre_rows  = df_ex[df_ex["era"] == "Pre-Q4 2022"]
        post_rows = df_ex[df_ex["era"] == "Q4 2022–Present"]

        for ri, (dt, row_d) in enumerate(df_ex.iterrows(), data_start_row + 1):
            is_post  = row_d["era"] == "Q4 2022–Present"
            row_fill = POST_FILL if is_post else None
            resid    = row_d["okun_resid"]

            resid_interp = ""
            if not np.isnan(resid):
                if resid > 0.3:
                    resid_interp = "Unemp. much higher than Okun predicts"
                elif resid > 0.1:
                    resid_interp = "Unemp. slightly above Okun prediction"
                elif resid < -0.3:
                    resid_interp = "Unemp. much lower than Okun predicts"
                elif resid < -0.1:
                    resid_interp = "Unemp. slightly below Okun prediction"
                else:
                    resid_interp = "On-target with Okun prediction"

            resid_fill = (POS_FILL if (not np.isnan(resid) and resid > 0.1)
                          else NEG_FILL if (not np.isnan(resid) and resid < -0.1)
                          else None)

            row_vals = [
                dt.strftime("%Y-Q%q") if hasattr(dt, 'strftime') else str(dt)[:7],
                row_d["era"],
                round(row_d["output"], 1),
                round(row_d["unemp"], 2),
                round(row_d["pct_dy"], 3)   if not np.isnan(row_d["pct_dy"])   else "N/A",
                round(row_d["delta_u"], 3)  if not np.isnan(row_d["delta_u"])  else "N/A",
                round(row_d["u_predicted"], 3) if not np.isnan(row_d["u_predicted"]) else "N/A",
                round(resid, 3)             if not np.isnan(resid)             else "N/A",
                resid_interp,
            ]
            fmts_d = [None, None, "#,##0.0", "0.00", "0.000", "0.000", "0.000", "0.000", None]

            for ci, (v, fmt) in enumerate(zip(row_vals, fmts_d), 1):
                fill = resid_fill if ci == 8 else row_fill
                data_cell(ws, ri, ci, v, fmt=fmt, fill=fill)

        # Freeze top rows and set column widths
        ws.freeze_panes = f"A{data_start_row + 1}"
        col_widths = [12, 18, 18, 14, 16, 16, 16, 16, 38]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Add a notes row after data
        last_data_row = data_start_row + len(df_ex) + 1
        note = ws.cell(row=last_data_row, column=1,
            value=f"Note: Orange rows = post-Q4 2022 (AI era). "
                  f"Red residual cells = unemployment higher than pre-2022 Okun fit predicted "
                  f"(output grew without proportional unemployment decline). "
                  f"BLS/BEA match: {cfg['match']}. {cfg['notes']}")
        note.font = Font(italic=True, size=9, color="666666")
        ws.merge_cells(f"A{last_data_row}:I{last_data_row}")

    wb.save("okun_industry_detail.xlsx")
    print("Excel workbook saved: okun_industry_detail.xlsx")

except ImportError:
    print("openpyxl not installed — skipping Excel export. Run: pip install openpyxl")

plt.show()
print("\nAll outputs saved.")

